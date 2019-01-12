import datetime
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.template import loader
from openpyxl import Workbook
from .models import Products

from static.etlapp import ScrapperRedux1

def index(request):
    products_list = Products.objects.order_by('-pub_date')
    widths_list = ScrapperRedux1.get_tire_widths()
    # template = loader.get_template('etlapp/index.html')
    context = {
        'products_list': products_list,
        'widths_list': widths_list,
    }
    # return HttpResponse(template.render(context, request))
    return render(request, 'etlapp/index.html', context)

def save_file(request):
    if request.method == "POST":
        data = {}        
        if request.POST['selected_product'] == ".TXT":
            product_id = Products.objects.get(id=request.POST['product_id'])
            txt_path = 'static/etlapp/files_from_database/dane.txt'
            generate_txt(product_id, txt_path)

            data = {
                "dfile": product_id.Name
            }
        else:
            generate_csv(Products.objects.all(), 'static/etlapp/files_from_database/oponeo.xls')
        return JsonResponse(data)

    return render(request, 'etlapp/index.html')

def generate_csv(Products, file_path):
    wb = Workbook()
    ws = wb.active
    row_count = 1

    ws['A' + str(row_count)] = "ID BD"
    ws['B' + str(row_count)] = "ID Oponeo"
    ws['C' + str(row_count)] = "Producent"
    ws['D' + str(row_count)] = "Nazwa"
    ws['E' + str(row_count)] = "Cena"
    ws['F' + str(row_count)] = "Typ pojazdu"
    ws['G' + str(row_count)] = "Sezon"
    ws['H' + str(row_count)] = "Rozmiar"
    ws['I' + str(row_count)] = "Homologacja"
    ws['J' + str(row_count)] = "Indeks prędkości"
    ws['K' + str(row_count)] = "Indeks nośności"
    ws['L' + str(row_count)] = "Etykieta UE"
    ws['M' + str(row_count)] = "Rok produkcji"
    ws['N' + str(row_count)] = "Kraj produkcji"
    ws['O' + str(row_count)] = "Gwarancja"
    ws['P' + str(row_count)] = "Inne"
    ws['Q' + str(row_count)] = "Data publikacji w bazie"
    row_count += 1

    for product in Products:
        ws['A' + str(row_count)] = product.id
        ws['B' + str(row_count)] = product.ProductID
        ws['C' + str(row_count)] = product.Manufacturer
        ws['D' + str(row_count)] = product.Name
        ws['E' + str(row_count)] = product.Price
        ws['F' + str(row_count)] = product.Car_type
        ws['G' + str(row_count)] = product.season
        ws['H' + str(row_count)] = product.size
        ws['I' + str(row_count)] = product.approval
        ws['J' + str(row_count)] = product.speed_index
        ws['K' + str(row_count)] = product.weight_index
        ws['L' + str(row_count)] = product.sound_index
        ws['M' + str(row_count)] = product.production_year
        ws['N' + str(row_count)] = product.country_of_origin
        ws['O' + str(row_count)] = product.guaranty
        ws['P' + str(row_count)] = product.other_info
        ws['Q' + str(row_count)] = product.pub_date
        row_count += 1

    wb.save(file_path)

def generate_txt(product, file_path):
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write("Produkt: " + str(product.Name) + '\n\n')
        file.write("Id bazy danych: " + str(product.id) + "\n")
        file.write("Id OPONEO: " + str(product.ProductID) + "\n")
        file.write("Producent: " + str(product.Manufacturer) + "\n")
        file.write("Cena: " + str(product.Price) + "\n")
        file.write("Typ pojazdu: " + str(product.Car_type) + "\n")
        file.write("Sezon: " + str(product.season) + "\n")
        file.write("Rozmiar: " + str(product.size) + "\n")
        file.write("Homologacja: " + str(product.approval) + "\n")
        file.write("Indeks prędkości: " + str(product.speed_index) + "\n")
        file.write("Indeks nośności: " + str(product.weight_index) + "\n")
        file.write("Etykieta UE: " + str(product.sound_index) + "\n")
        file.write("Rok produkcji: " + str(product.production_year) + "\n")
        file.write("Kraj produkcji: " + str(product.country_of_origin) + "\n")
        file.write("Gwarancja: " + str(product.guaranty) + "\n")
        file.write("Inne: " + str(product.other_info) + "\n")
        file.write("Data publikacji w bazie: " + str(product.pub_date) + "\n")

def clear_database(request):
    if request.method == "POST":
        summary = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " Baza danych została wyczyszczona. Usunięto " + str(len(Products.objects.all())) + " rekordów.\n"
        Products.objects.all().delete()

        data = {
            summary:summary,
        }

        with open('static/etlapp/logs/db_session_logs.txt', 'r', encoding='utf-8') as session_logs:
            with open('static/etlapp/logs/logs.txt', 'a', encoding='utf-8') as logs:
                logs.write(session_logs.read())

        with open('static/etlapp/logs/db_session_logs.txt', 'w', encoding='utf-8') as file:
                file.write(summary)

        with open('static/etlapp/logs/db_session_logs.txt', 'r', encoding='utf-8') as file:
            content = [x.strip('\n') for x in file.readlines()]

        test = ""
        for x in content:
            test = test + x + "\n"

        data = {
            summary:test,
        }

        return JsonResponse(data)

    return render(request, 'etlapp/index.html')

def refresh_table(request):
    #increment = int(request.GET['append_increment'])
    #increment_to = increment + 10
    products_list = Products.objects.all().order_by('-pub_date')
    return render(request, 'etlapp/refresh_table.html', {'products_list': products_list})

tires = list()
data_list = list()
date = str(datetime.datetime.now())
def dummy(request):
    global data_list
    global tires
    if request.method == "POST":
        data = {}
        summary = ""

        if request.POST['button_text'] == "ETL":
            width = request.POST['dropdown_id']
            time = ScrapperRedux1.extract(tires, width)
            summary += str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " Ekstrakcja zakończona. Sparsowano " + str(len(tires)) + " rekordów. Operacja zajęła " + str(int(time/60)) + " min " + str(int(time%60)) + " sekund.\n"
            data_list = ScrapperRedux1.transform(tires)
            summary += str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " Transformacja zakończona. Przygotowano " + str(len(data_list[0])) + " obiektów. Operacja zajęła " + str(int(data_list[1]/60)) + " min " + str(int(data_list[1]%60)) + " sekund.\n"
            time = ScrapperRedux1.load(data_list[0], date)
            summary += str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " Ładowanie zakończone. Operacja zajęła " + str(int(time/60)) + " min " + str(int(time%60)) + " sekund.\n"
            data_list = list()
            tires = list()

        elif request.POST['button_text'] == "EXTRACT":
            tire_widths = ScrapperRedux1.get_tire_widths()
            width = request.POST['dropdown_id']
            extract_time = ScrapperRedux1.extract(tires, width)
            summary += str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " Ekstrakcja zakończona. Sparsowano " + str(len(tires)) + " rekordów. Operacja zajęła " + str(int(extract_time/60)) + " min " + str(int(extract_time%60)) + " sekund.\n"

        elif request.POST['button_text'] == "TRANSFORM":
            data_list = ScrapperRedux1.transform(tires)
            summary += str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " Transformacja zakończona. Przygotowano " + str(len(data_list[0])) + " obiektów. Operacja zajęła " + str(int(data_list[1]/60)) + " min " + str(int(data_list[1]%60)) + " sekund.\n"

        else:
            time = ScrapperRedux1.load(data_list[0], date)
            summary += str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')) + " Ładowanie zakończone. Operacja zajęła " + str(int(time/60)) + " min " + str(int(time%60)) + " sekund.\n"
            data_list = list()
            tires = list()

        with open('static/etlapp/logs/db_session_logs.txt', 'a', encoding='utf-8') as file:
                file.write(summary)

        with open('static/etlapp/logs/db_session_logs.txt', 'r', encoding='utf-8') as file:
            content = [x.strip('\n') for x in file.readlines()]

        test = ""
        for x in content:
            test = test + x + "\n"

        data = {
            summary:test,
        }

        return JsonResponse(data)

    return render(request, 'etlapp/index.html')